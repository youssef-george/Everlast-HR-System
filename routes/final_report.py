from flask import Blueprint, render_template, request, jsonify, redirect, url_for, current_app, send_file
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime, date, timedelta
from models import User, DailyAttendance, LeaveRequest, PermissionRequest, AttendanceLog, PaidHoliday, db
from connection_manager import is_sync_running
from sqlalchemy import or_, and_, func
import logging
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
# Import shared calculation function
from report_helpers.report_calculations import calculate_unified_report_data, calculate_multiple_users_report_data
from helpers import format_hours_minutes

final_report_bp = Blueprint('final_report', __name__)

# Function moved to helpers/report_calculations.py for shared use

def _legacy_calculate_user_report_data(user, start_date, end_date):
    """
    Calculate report data for a single user using the exact same logic as the main final_report function.
    This ensures consistency between the web view and Excel export.
    """
    # Get attendance records for the user in the date range
    attendance_records = DailyAttendance.query.filter(
        DailyAttendance.user_id == user.id,
        DailyAttendance.date >= start_date,
        DailyAttendance.date <= end_date
    ).order_by(DailyAttendance.date.desc()).all()
    
    # Get leave requests for the user in the date range
    leave_requests = LeaveRequest.query.filter(
        LeaveRequest.user_id == user.id,
        LeaveRequest.start_date <= end_date,
        LeaveRequest.end_date >= start_date,
        LeaveRequest.status.in_(['approved', 'pending'])
    ).order_by(LeaveRequest.start_date.desc()).all()
    
    # Get permission requests for the user in the date range
    # Convert dates to datetime for comparison
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    permission_requests = PermissionRequest.query.filter(
        PermissionRequest.user_id == user.id,
        PermissionRequest.start_time <= end_datetime,
        PermissionRequest.end_time >= start_datetime,
        PermissionRequest.status.in_(['approved', 'pending'])
    ).order_by(PermissionRequest.start_time.desc()).all()
    
    # Calculate summary metrics using the EXACT same logic as the main function
    total_days = (end_date - start_date).days + 1
    
    # Calculate present and absent days more accurately
    present_days = 0
    absent_days = 0
    
    # Count actual days from attendance records
    for record in attendance_records:
        # A day is considered present if there's ANY log (check-in OR check-out) or if status is explicitly present
        if record.first_check_in or record.last_check_out or record.status in ['present', 'half-day', 'partial']:
            present_days += 1
        # Note: absent_days will be calculated later based on missing days in the date range
    
    # Calculate different types of leave days from actual leave requests
    annual_leave_days = 0  # This will now include both annual and sick leave
    unpaid_leave_days = 0
    paid_leave_days = 0
    
    for leave_request in leave_requests:
        
        # Calculate days within the date range
        leave_start = max(leave_request.start_date, start_date)
        leave_end = min(leave_request.end_date, end_date)
        if leave_start <= leave_end:
            days_count = (leave_end - leave_start).days + 1
            
            # Categorize by leave type
            if leave_request.leave_type:
                leave_type_name = leave_request.leave_type.name.lower()
                if 'annual' in leave_type_name or 'vacation' in leave_type_name or 'sick' in leave_type_name or 'illness' in leave_type_name:
                    # Combine annual and sick leave into one column
                    annual_leave_days += days_count
                elif 'unpaid' in leave_type_name:
                    unpaid_leave_days += days_count
                elif 'paid' in leave_type_name or 'holiday' in leave_type_name:
                    paid_leave_days += days_count
                else:
                    # Default to annual leave if type is unclear
                    annual_leave_days += days_count
            else:
                # Default to annual leave if no type specified
                annual_leave_days += days_count
    
    # Add paid holidays to paid leave days - only count holidays within the date range
    paid_holidays = PaidHoliday.query.filter(
        PaidHoliday.start_date <= end_date,
        db.or_(
            PaidHoliday.end_date.is_(None),
            PaidHoliday.end_date >= start_date
        )
    ).all()
    
    for paid_holiday in paid_holidays:
        # Calculate the actual overlap with the date range
        holiday_start = max(paid_holiday.start_date, start_date)
        holiday_end = min(paid_holiday.end_date or paid_holiday.start_date, end_date)
        
        if holiday_start <= holiday_end:
            # Count each day of the holiday that falls within the date range
            current_holiday_date = holiday_start
            while current_holiday_date <= holiday_end:
                # Check if this holiday day falls on a working day (Monday-Friday)
                if current_holiday_date.weekday() < 5:  # Monday = 0, Sunday = 6
                    paid_leave_days += 1
                current_holiday_date += timedelta(days=1)
    
    # Calculate permission hours from actual permission requests
    permission_hours = 0.0
    for permission_request in permission_requests:
        # Calculate hours within the date range
        perm_start = max(permission_request.start_time, datetime.combine(start_date, datetime.min.time()))
        perm_end = min(permission_request.end_time, datetime.combine(end_date, datetime.max.time()))
        if perm_start <= perm_end:
            duration = (perm_end - perm_start).total_seconds() / 3600  # Convert to hours
            permission_hours += duration
    
    # Calculate extra time (hours) using EXACT same logic as main final_report function
    extra_time_hours = 0.0
    
    # Process each day in the date range using Calendar Attendance Report logic
    current_date = start_date
    today = date.today()
    
    # Initialize processed_logs outside the loop (same as Calendar Attendance Report)
    processed_logs = {}
    
    while current_date <= end_date:
        # Initialize variables for this day
        status = 'Absent'
        check_in = None
        check_out = None
        hours_worked = 0.0
        
        # Check if before joining date
        if user.joining_date and current_date < user.joining_date:
            status = 'Not Yet Joined'
        elif current_date > today:
            status = 'Future Date'
        else:
            # Check for paid holidays first (same as Calendar Attendance Report)
            paid_holiday = PaidHoliday.query.filter(
                or_(
                    and_(PaidHoliday.holiday_type == 'day', PaidHoliday.start_date == current_date),
                    and_(PaidHoliday.holiday_type == 'range', 
                         PaidHoliday.start_date <= current_date, 
                         PaidHoliday.end_date >= current_date)
                )
            ).first()
            
            if paid_holiday:
                # Check if user has attendance logs
                start_datetime = datetime.combine(current_date, datetime.min.time())
                end_datetime = datetime.combine(current_date, datetime.max.time())
                today_logs = AttendanceLog.query.filter(
                    AttendanceLog.user_id == user.id,
                    AttendanceLog.timestamp.between(start_datetime, end_datetime)
                ).order_by(AttendanceLog.timestamp).all()
                
                if today_logs:
                    status = f"Present - {paid_holiday.description}"
                else:
                    status = paid_holiday.description
            else:
                # Get attendance logs for this date
                start_datetime = datetime.combine(current_date, datetime.min.time())
                end_datetime = datetime.combine(current_date, datetime.max.time())
                
                today_logs = AttendanceLog.query.filter(
                    AttendanceLog.user_id == user.id,
                    AttendanceLog.timestamp.between(start_datetime, end_datetime)
                ).order_by(AttendanceLog.timestamp).all()
            
            # Process logs using same logic as Calendar Attendance Report
            # Clear processed_logs for each date
            processed_logs.clear()
            for log in today_logs:
                if log.user_id not in processed_logs:
                    processed_logs[log.user_id] = {
                        'user': log.user,
                        'check_in': None,
                        'check_out': None,
                        'duration': None,
                        'status': 'absent',
                        'all_logs': []
                    }
                
                # Add log to all_logs
                processed_logs[log.user_id]['all_logs'].append(log)
            
            # Use dynamic attendance processing for each user
            for user_id_key, data in processed_logs.items():
                from routes.attendance import determine_attendance_type_dynamic
                attendance_result = determine_attendance_type_dynamic(data['all_logs'])
                
                data['check_in'] = attendance_result['check_in']
                data['check_out'] = attendance_result['check_out']
                data['is_incomplete'] = attendance_result['is_incomplete']
                
                # Set status based on dynamic result
                if data['check_in'] and data['check_out']:
                        data['status'] = 'present'
                        duration = data['check_out'].timestamp - data['check_in'].timestamp
                        data['duration'] = duration
                elif data['check_in'] and not data['check_out']:
                    data['status'] = 'present'  # Incomplete day still counts as present
                    data['duration'] = None
                else:
                    data['status'] = 'absent'
            
            # Check for day off (Friday/Saturday) - EXACT same as Calendar Attendance Report
            if current_date.weekday() in [4, 5]:  # Friday (4) or Saturday (5)
                if user.id in processed_logs and processed_logs[user.id]['status'] == 'present':
                    status = 'Day Off / Present'
                    check_in = processed_logs[user.id]['check_in'].timestamp if processed_logs[user.id]['check_in'] else None
                    check_out = processed_logs[user.id]['check_out'].timestamp if processed_logs[user.id]['check_out'] else None
                    is_incomplete = processed_logs[user.id].get('is_incomplete', False)
                    
                    # Only calculate extra time for complete days (not incomplete)
                    if check_in and check_out and not is_incomplete:
                        time_diff = check_out - check_in
                        hours_worked = time_diff.total_seconds() / 3600
                        # Calculate extra time: actual hours - required 9 hours
                        extra_time_hours += (hours_worked - 9)
                else:
                    status = 'Day Off'
            else:
                # Check for leave requests
                leave_request = LeaveRequest.query.filter(
                    LeaveRequest.user_id == user.id,
                    LeaveRequest.start_date <= current_date,
                    LeaveRequest.end_date >= current_date,
                    LeaveRequest.status == 'approved'
                ).first()
                
                # Check for permission requests
                permission_request = PermissionRequest.query.filter(
                    PermissionRequest.user_id == user.id,
                    func.date(PermissionRequest.start_time) == current_date,
                    PermissionRequest.status == 'approved'
                ).first()
                
                # Check if user has attendance logs
                has_attendance_logs = user.id in processed_logs
                
                if has_attendance_logs:
                    # User has attendance logs - prioritize showing attendance
                    user_data = processed_logs[user.id]
                    check_in = user_data['check_in'].timestamp if user_data['check_in'] else None
                    check_out = user_data['check_out'].timestamp if user_data['check_out'] else None
                    
                    if user_data['status'] == 'present':
                        # Calculate hours worked and extra time (only if no leave request and not incomplete)
                        is_incomplete = user_data.get('is_incomplete', False)
                        if check_in and check_out and not is_incomplete:
                            time_diff = check_out - check_in
                            hours_worked = time_diff.total_seconds() / 3600
                            # Calculate extra time only if no leave request and no permission request: actual hours - required 9 hours
                            if not leave_request and not permission_request:
                                extra_time_hours += (hours_worked - 9)
                    elif user_data['status'] == 'in_office':
                        # Calculate hours worked and extra time for in_office status (only if no leave request and not incomplete)
                        is_incomplete = user_data.get('is_incomplete', False)
                        if check_in and check_out and not is_incomplete:
                            time_diff = check_out - check_in
                            hours_worked = time_diff.total_seconds() / 3600
                            # Calculate extra time only if no leave request and no permission request: actual hours - required 9 hours
                            if not leave_request and not permission_request:
                                extra_time_hours += (hours_worked - 9)
                    else:
                        # Calculate hours worked and extra time for other statuses (only if no leave request and not incomplete)
                        is_incomplete = user_data.get('is_incomplete', False)
                        if check_in and check_out and not is_incomplete:
                            time_diff = check_out - check_in
                            hours_worked = time_diff.total_seconds() / 3600
                            # Calculate extra time only if no leave request and no permission request: actual hours - required 9 hours
                            if not leave_request and not permission_request:
                                extra_time_hours += (hours_worked - 9)
                else:
                    # User has no attendance logs - check for leave/permission/paid holiday
                    if not leave_request and not permission_request and not paid_holiday:
                        # Only count as absent if it's not a future date
                        if current_date <= today:
                            status = 'Absent'
                            # Don't calculate negative working hours for absent days
        
        current_date += timedelta(days=1)
    
    # Calculate working days (Monday to Thursday) and Friday+Saturday days
    working_days = 0
    total_friday_saturday_days = 0
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 4:  # Monday = 0, Tuesday = 1, Wednesday = 2, Thursday = 3
            working_days += 1
        elif current_date.weekday() in [4, 5]:  # Friday = 4, Saturday = 5
            total_friday_saturday_days += 1
        # Sunday = 6 is not counted in either
        current_date += timedelta(days=1)
    
    # Calculate total leave days
    total_leave_days = annual_leave_days + unpaid_leave_days + paid_leave_days
    total_attendance_days = present_days + total_leave_days
    
    # Calculate actual day off days: total Friday+Saturday minus days worked/taken as leave
    # Count present days, annual leave, and paid leave that fall on Friday/Saturday
    friday_saturday_used_days = 0
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() in [4, 5]:  # Friday = 4, Saturday = 5
            # Check if this Friday/Saturday was used for work or leave
            daily_record = None
            if attendance_records:
                daily_record = next((r for r in attendance_records if r.date == current_date), None)
            
            # Check for leave requests on this date
            leave_on_date = any(
                lr.start_date <= current_date <= lr.end_date and lr.status == 'approved'
                for lr in leave_requests
            )
            
            # Check for paid holidays on this date
            paid_holiday_on_date = any(
                (ph.holiday_type == 'day' and ph.start_date == current_date) or
                (ph.holiday_type == 'range' and ph.start_date <= current_date <= ph.end_date)
                for ph in paid_holidays
            )
            
            # If there's attendance, leave, or paid holiday on this Friday/Saturday, count it as used
            if (daily_record and daily_record.status == 'present') or leave_on_date or paid_holiday_on_date:
                friday_saturday_used_days += 1
        
        current_date += timedelta(days=1)
    
    # Day off days = Total Friday+Saturday days - Used Friday+Saturday days
    day_off_days = total_friday_saturday_days - friday_saturday_used_days
    day_off_days = max(0, day_off_days)  # Ensure it's not negative
    
    # Calculate absent days properly: count each working day in the range
    absent_days = 0
    current_date = start_date
    today = date.today()
    
    while current_date <= end_date:
        # Count working days: Monday to Thursday (0-3) and Sunday (6)
        # Friday (4) and Saturday (5) are day off and not counted as absent
        if current_date.weekday() < 4 or current_date.weekday() == 6:  # Monday-Thursday (0-3) or Sunday (6)
            # Only count if date is on or after joining date and not in the future
            if (not user.joining_date or current_date >= user.joining_date) and current_date <= today:
                # Check if there's any attendance record for this date (with actual attendance data)
                has_attendance = any(
                    record.date == current_date and 
                    (record.first_check_in or record.last_check_out or record.status in ['present', 'half-day', 'partial', 'in_office'])
                    for record in attendance_records
                )
                
                # Also check for raw attendance logs (in case DailyAttendance record doesn't exist yet)
                if not has_attendance:
                    start_datetime = datetime.combine(current_date, datetime.min.time())
                    end_datetime = datetime.combine(current_date, datetime.max.time())
                    today_logs = AttendanceLog.query.filter(
                        AttendanceLog.user_id == user.id,
                        AttendanceLog.timestamp.between(start_datetime, end_datetime)
                    ).count()
                    if today_logs > 0:
                        has_attendance = True
                
                # Check for leave requests on this date (both approved and pending should exclude from absent)
                has_leave = any(
                    lr.start_date <= current_date <= lr.end_date and lr.status in ['approved', 'pending']
                    for lr in leave_requests
                )
                
                # Check for paid holidays on this date
                has_paid_holiday = any(
                    (ph.holiday_type == 'day' and ph.start_date == current_date) or
                    (ph.holiday_type == 'range' and ph.start_date <= current_date <= ph.end_date)
                    for ph in paid_holidays
                )
                
                # Check for permission requests on this date (both approved and pending should exclude from absent)
                has_permission = any(
                    pr.start_time.date() <= current_date <= pr.end_time.date() and pr.status in ['approved', 'pending']
                    for pr in permission_requests
                )
                
                # Count as absent only if no attendance, leave, holiday, or permission
                if not has_attendance and not has_leave and not has_paid_holiday and not has_permission:
                    absent_days += 1
        
        current_date += timedelta(days=1)
    
    # Calculate incomplete days based on FIXED business rules:
    # - ONLY 1 log entry total: Mark as incomplete
    # - Multiple log entries (even with same timestamp): Complete day
    # - No logs: Absent (not incomplete)
    # FIXED: Days with 10+ logs are now correctly marked as complete
    incomplete_days = 0
    if attendance_records:
        incomplete_days = sum(1 for record in attendance_records if record.is_incomplete_day)
    
    # Calculate total working days as: day off + present + annual + paid
    total_working_days = day_off_days + present_days + annual_leave_days + paid_leave_days
    
    # Calculate attendance percentage based on total working days
    attendance_percentage = (present_days / total_working_days * 100) if total_working_days > 0 else 0
    
    summary_metrics = {
        'total_days': total_days,
        'total_working_days': total_working_days,  # Day Off + Present + Annual + Paid
        'present_days': present_days,
        'absent_days': absent_days,
        'annual_leave_days': annual_leave_days,  # Now includes both annual and sick leave
        'unpaid_leave_days': unpaid_leave_days,
        'paid_leave_days': paid_leave_days,
        'permission_hours': round(permission_hours, 2),
        'day_off_days': day_off_days,
        'incomplete_days': incomplete_days,  # Days with only one log
        'attendance_percentage': round(attendance_percentage, 1),
        'extra_time_hours': round(extra_time_hours, 1)
    }
    
    return {
        'user': user,
        'summary_metrics': summary_metrics,
        'attendance_records': attendance_records,
        'leave_requests': leave_requests,
        'permission_requests': permission_requests
    }

def role_required(roles):
    """Decorator to require specific roles"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth.login'))
            if current_user.role not in roles:
                return redirect(url_for('dashboard.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def ensure_attendance_logs_processed(start_date, end_date):
    """Ensure all attendance logs in the date range are processed into DailyAttendance records"""
    from models import AttendanceLog, User, DailyAttendance
    from routes.attendance import process_daily_attendance
    from sqlalchemy import func, and_
    import logging
    
    try:
        # Get all attendance logs in the date range
        attendance_logs = AttendanceLog.query.filter(
            func.date(AttendanceLog.timestamp) >= start_date,
            func.date(AttendanceLog.timestamp) <= end_date
        ).all()
        
        # Group by user and date (convert to Python date objects)
        user_dates = set()
        for log in attendance_logs:
            log_date = log.timestamp.date()  # Convert to Python date object
            user_dates.add((log.user_id, log_date))
        
        # Check which ones already have DailyAttendance records
        user_dates_to_process = []
        for user_id, log_date in user_dates:
            existing_record = DailyAttendance.query.filter_by(
                user_id=user_id,
                date=log_date
            ).first()
            
            if not existing_record:
                user_dates_to_process.append((user_id, log_date))
        
        # Only process user-date combinations that don't have DailyAttendance records
        processed_count = 0
        for user_id, log_date in user_dates_to_process:
            try:
                # Ensure we're using a fresh session for each processing
                db.session.rollback()  # Clear any pending transactions
                process_daily_attendance(user_id, log_date)
                db.session.commit()  # Commit each one individually
                processed_count += 1
            except Exception as e:
                logging.error(f"Error processing daily attendance for user {user_id} on {log_date}: {str(e)}")
                db.session.rollback()  # Rollback on error
                continue
        
        logging.info(f"Processed {processed_count} new user-date combinations for attendance logs")
        
    except Exception as e:
        logging.error(f"Error in ensure_attendance_logs_processed: {str(e)}")
        db.session.rollback()
        raise

@final_report_bp.route('/final-report')
@login_required
@role_required(['admin', 'product_owner', 'manager', 'employee'])
def final_report():
    """Final Report - Admin, Product Owner, Manager, and Employee attendance report with auto-fetch and duplicate removal"""
    
    try:
        # Clean up orphaned paid holiday records before processing
        from routes.attendance import cleanup_orphaned_paid_holiday_records
        with current_app.app_context():
            cleanup_orphaned_paid_holiday_records()
    except Exception as e:
        logging.error(f'Error cleaning up paid holiday records: {str(e)}')
    
    # Auto-sync data from devices
    if not is_sync_running():
        try:
            from routes.attendance import sync_attendance_task
            def sync_task():
                try:
                    with current_app.app_context():
                        sync_attendance_task(full_sync=True)
                except Exception as e:
                    logging.error(f'Error auto-syncing data on final report page load: {str(e)}')
            
            sync_thread = threading.Thread(target=sync_task, daemon=True)
            sync_thread.start()
        except Exception as e:
            logging.error(f'Error starting sync thread for final report: {str(e)}')
    else:
        logging.info('Skipping sync on final report page load - another sync is already running')
    
    # Get users based on role
    if current_user.role == 'employee':
        # Employees can only see their own data
        users = [current_user] if current_user.status == 'active' else []
    elif current_user.role == 'manager':
        # Managers see their employees AND themselves
        from helpers import get_employees_for_manager
        team_members = get_employees_for_manager(current_user.id)
        # Include manager themselves
        users = [current_user] + list(team_members)
        # Filter to active users only
        users = [u for u in users if u.status == 'active' and 
                 not u.first_name.startswith('User') and 
                 not u.first_name.startswith('NN-') and
                 u.first_name != '' and u.last_name != '']
    else:
        # Admin and Product Owner see all active users
        users = User.query.filter(
            User.status == 'active',
            ~User.first_name.like('User%'),  # Exclude generic test users
            ~User.first_name.like('NN-%'),   # Exclude numbered test users
            User.first_name != '',           # Exclude empty names
            User.last_name != ''             # Exclude users without last names
        ).all()
    
    # Sort users by fingerprint number: Not Assigned (None/empty) first, then numeric low to high
    def get_fingerprint_sort_key(user):
        fp = user.fingerprint_number
        if not fp or fp.strip() == '':
            return (0, 0)  # Not Assigned comes first
        try:
            fp_int = int(fp.strip())
            return (1, fp_int)  # Numeric sorting
        except (ValueError, TypeError):
            return (2, 0)  # Invalid/non-numeric comes last
    
    users = sorted(users, key=get_fingerprint_sort_key)
    
    # Get date range from query parameters (no default - user must choose)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    
    # Only process if dates are provided
    if not start_date_str or not end_date_str:
        # Return empty report if no dates provided
        return render_template('final_report/index.html', 
                             users=users, 
                             start_date=None, 
                             end_date=None, 
                             all_user_reports=[])
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Ensure all attendance logs are processed into DailyAttendance records (optional, for performance)
    process_logs = request.args.get('process_logs', 'true').lower() == 'true'
    if process_logs:
        try:
            ensure_attendance_logs_processed(start_date, end_date)
        except Exception as e:
            logging.error(f"Error processing attendance logs: {str(e)}")
            # Rollback any pending transactions and continue with report generation
            db.session.rollback()
    
    # Filter users if specific users are selected
    if user_ids:
        # For employees, ensure they can only view themselves
        if current_user.role == 'employee':
            user_ids = [current_user.id] if current_user.id in user_ids else []
        # For managers, ensure they can only view their own employees and themselves
        elif current_user.role == 'manager':
            manager_employee_ids = [u.id for u in users]
            # Always allow manager to view themselves
            if current_user.id not in manager_employee_ids:
                manager_employee_ids.append(current_user.id)
            user_ids = [uid for uid in user_ids if uid in manager_employee_ids]
        users = [user for user in users if user.id in user_ids]
    
    # Sort users by fingerprint number: Not Assigned (None/empty) first, then numeric low to high
    def get_fingerprint_sort_key(user):
        fp = user.fingerprint_number
        if not fp or fp.strip() == '':
            return (0, 0)  # Not Assigned comes first
        try:
            fp_int = int(fp.strip())
            return (1, fp_int)  # Numeric sorting
        except (ValueError, TypeError):
            return (2, 0)  # Invalid/non-numeric comes last
    
    users = sorted(users, key=get_fingerprint_sort_key)
    
    # Generate report data using unified calculation logic
    all_user_reports = []
    
    for user in users:
        # Use the unified calculation function to ensure exact same logic across all reports
        user_report = calculate_unified_report_data(user, start_date, end_date)
    
        # Use the unified calculation result
        summary_metrics = user_report.summary_metrics
    
        # Create report data with duplicate removal (using unified function results)
        report_data = []

        # Add attendance records
        for record in user_report.attendance_records:
            # Use pre-calculated values from unified calculation function
            # This ensures consistency with all other reports
            hours_worked = getattr(record, 'hours_worked', 0.0)
            extra_time = getattr(record, 'extra_time', 0.0)
            
            # If not set by unified function, calculate using same logic
            if hours_worked == 0.0 and record.first_check_in and record.last_check_out:
                time_diff = record.last_check_out - record.first_check_in
                hours_worked = time_diff.total_seconds() / 3600
                extra_time = hours_worked - 9  # Standard working hours is 9
            
            report_data.append({
                'date': record.date,
                'status': record.status,
                'check_in': record.first_check_in,
                'check_out': record.last_check_out,
                'hours_worked': hours_worked,
                'extra_time': extra_time
            })

        # Add leave requests
        for leave_request in user_report.leave_requests:
            # Calculate days within the date range
            leave_start = max(leave_request.start_date, start_date)
            leave_end = min(leave_request.end_date, end_date)
            if leave_start <= leave_end:
                # Add each day of the leave
                current_leave_date = leave_start
                while current_leave_date <= leave_end:
                    report_data.append({
                        'date': current_leave_date,
                        'status': 'Leave',
                        'check_in': None,
                        'check_out': None,
                        'hours_worked': 0,
                        'extra_time': 0,
                        'leave_type': leave_request.leave_type.name if leave_request.leave_type else 'Unknown',
                        'leave_reason': leave_request.reason
                    })
                    current_leave_date += timedelta(days=1)

        # Add permission requests
        for permission_request in user_report.permission_requests:
            # Calculate days within the date range
            perm_start = max(permission_request.start_time, datetime.combine(start_date, datetime.min.time()))
            perm_end = min(permission_request.end_time, datetime.combine(end_date, datetime.max.time()))
            if perm_start <= perm_end:
                # Add each day of the permission
                current_perm_date = perm_start.date()
                while current_perm_date <= perm_end.date():
                    if current_perm_date >= start_date and current_perm_date <= end_date:
                        report_data.append({
                            'date': current_perm_date,
                            'status': 'Permission',
                            'check_in': None,
                            'check_out': None,
                            'hours_worked': 0,
                            'extra_time': 0,
                            'permission_reason': permission_request.reason,
                            'permission_start': permission_request.start_time,
                            'permission_end': permission_request.end_time
                        })
                    current_perm_date += timedelta(days=1)

        # No sorting needed for summary view

        all_user_reports.append({
            'user': user,
            'summary_metrics': summary_metrics,
            'report_data': report_data,
            'leave_requests': user_report.leave_requests, # Added missing
            'permission_requests': user_report.permission_requests # Added missing
        })

    try:
        return render_template('final_report/index.html', 
                             users=users, 
                             start_date=start_date, 
                             end_date=end_date, 
                             all_user_reports=all_user_reports)
    except Exception as e:
        logging.error(f'Error rendering final report template: {str(e)}')
        return render_template('final_report/index.html', 
                             users=[], 
                             start_date=start_date, 
                             end_date=end_date, 
                             all_user_reports=[],
                             error_message=f"Error loading report: {str(e)}")

@final_report_bp.route('/final-report/export')
@login_required
@role_required(['admin', 'product_owner', 'manager', 'employee'])
def export_final_report():
    """Export final report to Excel"""
    
    # Get the same parameters as the final report
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    
    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Start date and end date are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    # Get users for export based on role
    if current_user.role == 'employee':
        # Employees can only export their own data
        users = [current_user] if current_user.status == 'active' else []
    elif current_user.role == 'manager':
        # Managers can export their employees AND themselves
        from helpers import get_employees_for_manager
        team_members = get_employees_for_manager(current_user.id)
        # Include manager themselves
        users = [current_user] + list(team_members)
        # Filter to active users only
        users = [u for u in users if u.status == 'active' and 
                 not u.first_name.startswith('User') and 
                 not u.first_name.startswith('NN-') and
                 u.first_name != '' and u.last_name != '']
    else:
        # Admin and Product Owner can export all users
        users = User.query.filter(
            User.status == 'active',
            ~User.first_name.like('User%'),  # Exclude generic test users
            ~User.first_name.like('NN-%'),   # Exclude numbered test users
            User.first_name != '',           # Exclude empty names
            User.last_name != ''             # Exclude users without last names
        ).all()
    
    # Filter users if specific users are selected
    if user_ids:
        # For employees, ensure they can only export themselves
        if current_user.role == 'employee':
            user_ids = [current_user.id] if current_user.id in user_ids else []
        # For managers, ensure they can only export their own employees and themselves
        elif current_user.role == 'manager':
            manager_employee_ids = [u.id for u in users]
            # Always allow manager to export themselves
            if current_user.id not in manager_employee_ids:
                manager_employee_ids.append(current_user.id)
            user_ids = [uid for uid in user_ids if uid in manager_employee_ids]
        users = [user for user in users if user.id in user_ids]
    
    # Sort users by fingerprint number: Not Assigned (None/empty) first, then numeric low to high
    def get_fingerprint_sort_key(user):
        fp = user.fingerprint_number
        if not fp or fp.strip() == '':
            return (0, 0)  # Not Assigned comes first
        try:
            fp_int = int(fp.strip())
            return (1, fp_int)  # Numeric sorting
        except (ValueError, TypeError):
            return (2, 0)  # Invalid/non-numeric comes last
    
    users = sorted(users, key=get_fingerprint_sort_key)
    
    # Generate report data using the EXACT SAME logic as the main route
    all_user_reports = []
    
    for user in users:
        # Use the unified calculation function to ensure exact same logic as web view
        user_report = calculate_unified_report_data(user, start_date, end_date)
        all_user_reports.append(user_report)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Name", "Fingerprint Number", "Department", "Total Days", 
        "Total Working Days", "Day Off", "Present Days", "Absent Days", 
        "Annual Leave", "Paid Leave", "Permission Hours", "Incomplete Days", "Extra Time (hours)"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Write data
    for row, user_report in enumerate(all_user_reports, 2):
        user = user_report.user
        metrics = user_report.summary_metrics
        
        ws.cell(row=row, column=1, value=user.get_full_name()).border = border
        ws.cell(row=row, column=2, value=user.fingerprint_number or 'N/A').border = border
        ws.cell(row=row, column=3, value=user.department.department_name if user.department else 'No Department').border = border
        ws.cell(row=row, column=4, value=metrics.total_days).border = border
        ws.cell(row=row, column=5, value=metrics.total_working_days).border = border
        ws.cell(row=row, column=6, value=metrics.day_off_days).border = border
        ws.cell(row=row, column=7, value=metrics.present_days).border = border
        ws.cell(row=row, column=8, value=metrics.absent_days).border = border
        ws.cell(row=row, column=9, value=metrics.annual_leave_days).border = border
        ws.cell(row=row, column=10, value=metrics.paid_leave_days).border = border
        ws.cell(row=row, column=11, value=metrics.permission_hours).border = border
        ws.cell(row=row, column=12, value=metrics.incomplete_days).border = border
        # Format extra time as hours and minutes for Excel
        extra_time_formatted = format_hours_minutes(metrics.extra_time_hours)
        ws.cell(row=row, column=13, value=extra_time_formatted).border = border
    
    # Add summary row
    if all_user_reports:
        summary_row = len(all_user_reports) + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value="").border = border
        ws.cell(row=summary_row, column=3, value="").border = border
        ws.cell(row=summary_row, column=4, value=sum(r.summary_metrics.total_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=5, value=sum(r.summary_metrics.total_working_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=6, value=sum(r.summary_metrics.day_off_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=7, value=sum(r.summary_metrics.present_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=8, value=sum(r.summary_metrics.absent_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=9, value=sum(r.summary_metrics.annual_leave_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=10, value=sum(r.summary_metrics.paid_leave_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=11, value=sum(r.summary_metrics.permission_hours for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=12, value=sum(r.summary_metrics.incomplete_days for r in all_user_reports)).border = border
        # Format total extra time as hours and minutes
        total_extra_time = sum(r.summary_metrics.extra_time_hours for r in all_user_reports)
        total_extra_time_formatted = format_hours_minutes(total_extra_time)
        ws.cell(row=summary_row, column=13, value=total_extra_time_formatted).border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"Final_Report_{start_date_str}_to_{end_date_str}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def role_required(roles):
    """Decorator to check if user has required role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth.login'))
            if current_user.role not in roles:
                return redirect(url_for('dashboard.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@final_report_bp.route('/detailed-attendance-report')
@login_required
@role_required(['admin', 'director', 'support', 'product_owner', 'manager'])
def detailed_attendance_report():
    """Detailed Attendance Report - Admin, Director, Support, and Product Owner attendance report with expandable employee logs"""
    
    try:
        # Clean up orphaned paid holiday records before processing
        from routes.attendance import cleanup_orphaned_paid_holiday_records
        with current_app.app_context():
            cleanup_orphaned_paid_holiday_records()
    except Exception as e:
        logging.error(f'Error cleaning up paid holiday records: {str(e)}')
    
    # Auto-sync data from devices
    if not is_sync_running():
        try:
            from routes.attendance import sync_attendance_task
            def sync_task():
                try:
                    with current_app.app_context():
                        sync_attendance_task(full_sync=True)
                except Exception as e:
                    logging.error(f'Error auto-syncing data on detailed attendance report page load: {str(e)}')
            
            sync_thread = threading.Thread(target=sync_task, daemon=True)
            sync_thread.start()
        except Exception as e:
            logging.error(f'Error starting sync thread for detailed attendance report: {str(e)}')
    else:
        logging.info('Skipping sync on detailed attendance report page load - another sync is already running')
    
    # Get users based on role
    if current_user.role == 'manager':
        # Managers see their employees AND themselves
        from helpers import get_employees_for_manager
        team_members = get_employees_for_manager(current_user.id)
        # Include manager themselves
        users = [current_user] + list(team_members)
        # Filter to active users only
        users = [u for u in users if u.status == 'active' and 
                 not u.first_name.startswith('User') and 
                 not u.first_name.startswith('NN-') and
                 u.first_name != '' and u.last_name != '']
    else:
        # Admin, Director, Support, Product Owner see all active users
        users = User.query.filter(
            User.status == 'active',
            ~User.first_name.like('User%'),  # Exclude generic test users
            ~User.first_name.like('NN-%'),   # Exclude numbered test users
            User.first_name != '',           # Exclude empty names
            User.last_name != ''             # Exclude users without last names
        ).all()
    
    # Sort users by fingerprint number: Not Assigned (None/empty) first, then numeric low to high
    def get_fingerprint_sort_key(user):
        fp = user.fingerprint_number
        if not fp or fp.strip() == '':
            return (0, 0)  # Not Assigned comes first
        try:
            fp_int = int(fp.strip())
            return (1, fp_int)  # Numeric sorting
        except (ValueError, TypeError):
            return (2, 0)  # Invalid/non-numeric comes last
    
    users = sorted(users, key=get_fingerprint_sort_key)
    
    # Get date range from query parameters (no default - user must choose)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    
    # Only process if dates are provided
    if not start_date_str or not end_date_str:
        # Return empty report if no dates provided
        return render_template('final_report/detailed_report.html', 
                             users=users, 
                             start_date=None, 
                             end_date=None, 
                             all_user_reports=[])
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Ensure all attendance logs are processed into DailyAttendance records (optional, for performance)
    process_logs = request.args.get('process_logs', 'true').lower() == 'true'
    if process_logs:
        try:
            ensure_attendance_logs_processed(start_date, end_date)
        except Exception as e:
            logging.error(f"Error processing attendance logs: {str(e)}")
            # Rollback any pending transactions and continue with report generation
            db.session.rollback()
    
    # Filter users if specific users are selected
    if user_ids:
        # For managers, ensure they can only view their own employees and themselves
        if current_user.role == 'manager':
            manager_employee_ids = [u.id for u in users]
            # Always allow manager to view themselves
            if current_user.id not in manager_employee_ids:
                manager_employee_ids.append(current_user.id)
            user_ids = [uid for uid in user_ids if uid in manager_employee_ids]
        users = [user for user in users if user.id in user_ids]
    
    # Generate report data with duplicate removal - using exact same logic as final_report
    all_user_reports = []
    
    for user in users:
        # Use the unified calculation function to ensure exact same logic as Final Report web view
        user_report = calculate_unified_report_data(user, start_date, end_date)
        
        all_user_reports.append({
            'user': user,
            'summary_metrics': user_report.summary_metrics,
            'report_data': user_report  # Pass the entire user_report object
        })
    
    try:
        return render_template('final_report/detailed_report.html', 
                             users=users, 
                             start_date=start_date, 
                             end_date=end_date, 
                             all_user_reports=all_user_reports)
    except Exception as e:
        logging.error(f'Error rendering detailed attendance report template: {str(e)}')
        return render_template('final_report/detailed_report.html', 
                             users=[], 
                             start_date=start_date, 
                             end_date=end_date, 
                             all_user_reports=[],
                             error_message=f"Error loading report: {str(e)}")

@final_report_bp.route('/detailed-attendance-report/employee-logs/<int:user_id>', methods=['GET'])
@login_required
@role_required(['admin', 'director', 'support', 'product_owner', 'manager'])
def get_employee_logs(user_id):
    """API endpoint to fetch detailed attendance data for a specific employee for all days in date range"""
    logging.info(f"User {user_id} requested detailed report from {request.args.get('start_date')} to {request.args.get('end_date')}")

    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Start date and end date are required'}), 400

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Get user
        user = User.query.get_or_404(user_id)
        
        # For managers, ensure they can only view their own employees and themselves
        if current_user.role == 'manager':
            from helpers import get_employees_for_manager
            team_members = get_employees_for_manager(current_user.id)
            manager_employee_ids = [u.id for u in team_members] + [current_user.id]
            if user_id not in manager_employee_ids:
                return jsonify({'error': 'Access denied. You can only view reports for yourself and your team members.'}), 403
        
        logging.info(f"Processing report for user: {user.get_full_name()} (ID: {user.id})")

        # Use the unified calculation logic to get comprehensive data
        from report_helpers.report_calculations import calculate_unified_report_data
        user_report = calculate_unified_report_data(user, start_date, end_date)
        logging.info(f"Unified report data calculated. Attendance records: {len(user_report.attendance_records)}, Leave requests: {len(user_report.leave_requests)}, Permission requests: {len(user_report.permission_requests)}")

        # Get raw attendance logs for detailed view
        attendance_logs = AttendanceLog.query.filter(
            AttendanceLog.user_id == user.id,
            func.date(AttendanceLog.timestamp) >= start_date,
            func.date(AttendanceLog.timestamp) <= end_date
        ).order_by(AttendanceLog.timestamp.desc()).all()
        logging.info(f"Fetched {len(attendance_logs)} raw attendance logs for user {user.id}")

        # Group logs by date for detailed view
        logs_by_date = {}
        for log in attendance_logs:
            log_date = log.timestamp.date()
            if log_date not in logs_by_date:
                logs_by_date[log_date] = []
            logs_by_date[log_date].append({
                'id': log.id,
                'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'date': log.timestamp.date().strftime('%Y-%m-%d'),
                'time': log.timestamp.strftime('%H:%M:%S'),
                'scan_type': log.scan_type,
                'duration': log.format_duration() if hasattr(log, 'format_duration') and log.format_duration() else '-',
                'device_ip': log.device_ip,
                'is_extra_scan': log.is_extra_scan
            })
        
        # Sort logs within each date chronologically (oldest first)
        for date_key in logs_by_date:
            logs_by_date[date_key].sort(key=lambda x: x['time'])
        logging.info(f"Grouped raw logs into {len(logs_by_date)} unique dates.")

        # Format daily data for JSON response
        daily_data = []

        # Create daily records from the unified report data
        current_date = start_date
        while current_date <= end_date:
            try:
                logging.debug(f"Processing date: {current_date}")

                # Find attendance record for this date
                attendance_record = next(
                    (record for record in user_report.attendance_records if record.date == current_date), 
                    None
                )
                logging.debug(f"  Attendance record found: {attendance_record is not None}")

                # Find leave request for this date
                leave_request = next(
                    (leave_req for leave_req in user_report.leave_requests 
                     if leave_req.start_date <= current_date <= leave_req.end_date), 
                    None
                )
                logging.debug(f"  Leave request found: {leave_request is not None}")

                # Find permission request for this date
                permission_request = next(
                    (perm_req for perm_req in user_report.permission_requests 
                     if perm_req.start_time and perm_req.end_time and \
                        perm_req.start_time.date() <= current_date <= perm_req.end_time.date()), 
                    None
                )
                logging.debug(f"  Permission request found: {permission_request is not None}")

                # Initialize variables
                hours_worked = 0.0
                extra_time = 0.0
                check_in = None
                check_out = None
                status = 'Absent'
                day_of_week = current_date.strftime('%A')
                
                # Get logs for this date
                daily_logs = logs_by_date.get(current_date, [])
                logging.debug(f"  Daily logs count: {len(daily_logs)}")

                # Check if before joining date
                if user.joining_date and current_date < user.joining_date:
                    status = 'Not Yet Joined'
                # Don't show attendance data for future dates
                elif current_date > datetime.today().date():
                    status = 'Future Date'
                else:
                    # Check for paid holidays first
                    paid_holiday = PaidHoliday.query.filter(
                        or_(
                            and_(PaidHoliday.holiday_type == 'day', PaidHoliday.start_date == current_date),
                            and_(PaidHoliday.holiday_type == 'range', 
                                 PaidHoliday.start_date <= current_date, 
                                 PaidHoliday.end_date >= current_date)
                            )
                        ).first()
                    
                    if paid_holiday:
                        if attendance_record and (attendance_record.first_check_in or attendance_record.last_check_out):
                            status = f"Present - {paid_holiday.description}"
                            # Use pre-calculated values from unified calculation
                            if attendance_record:
                                hours_worked = getattr(attendance_record, 'hours_worked', 0.0)
                                extra_time = getattr(attendance_record, 'extra_time', 0.0)
                                check_in = attendance_record.first_check_in
                                check_out = attendance_record.last_check_out
                                
                                # FIX: Always use the last log as check-out when there are 2+ logs
                                if len(daily_logs) >= 2:
                                    # Sort logs by timestamp and use the last log's timestamp as check-out
                                    sorted_logs = sorted(daily_logs, key=lambda x: x['timestamp'])
                                    last_log_timestamp = sorted_logs[-1]['timestamp']
                                    last_log_dt = datetime.strptime(last_log_timestamp, '%Y-%m-%d %H:%M:%S')
                                    check_out = last_log_dt
                        else:
                            status = paid_holiday.description
                    elif attendance_record and (attendance_record.first_check_in or attendance_record.last_check_out):
                        # User has attendance logs - use pre-calculated values from unified calculation
                        check_in = attendance_record.first_check_in
                        check_out = attendance_record.last_check_out
                        
                        # FIX: Always use the last log as check-out when there are 2+ logs
                        if len(daily_logs) >= 2:
                            # Sort logs by timestamp and use the last log's timestamp as check-out
                            sorted_logs = sorted(daily_logs, key=lambda x: x['timestamp'])
                            last_log_timestamp = sorted_logs[-1]['timestamp']
                            last_log_dt = datetime.strptime(last_log_timestamp, '%Y-%m-%d %H:%M:%S')
                            check_out = last_log_dt
                        
                        hours_worked = getattr(attendance_record, 'hours_worked', 0.0)
                        extra_time = getattr(attendance_record, 'extra_time', 0.0)
                        
                        # Only recalculate if values are not set (fallback)
                        if hours_worked == 0.0 and check_in and check_out:
                            time_diff = check_out - check_in
                            hours_worked = time_diff.total_seconds() / 3600
                            extra_time = hours_worked - 9  # Standard working hours is 9
                        
                        # Determine status based on day and leave/permission
                        if current_date.weekday() in [4, 5]:  # Friday/Saturday
                            status = 'Day Off / Present'
                        elif leave_request:
                            leave_type_name = leave_request.leave_type.name if leave_request.leave_type else 'Leave Request'
                            status = f"Present / {leave_type_name}"
                        elif permission_request:
                            status = 'Present / Permission'
                        else:
                            status = attendance_record.status if attendance_record.status else 'Present'
                    elif daily_logs and len(daily_logs) > 0:
                        # Has logs but no attendance record - calculate from logs
                        first_log = daily_logs[0]['timestamp']
                        last_log = daily_logs[-1]['timestamp']
                        first_log_dt = datetime.strptime(first_log, '%Y-%m-%d %H:%M:%S')
                        last_log_dt = datetime.strptime(last_log, '%Y-%m-%d %H:%M:%S')
                        time_diff = last_log_dt - first_log_dt
                        hours_worked = time_diff.total_seconds() / 3600
                        extra_time = hours_worked - 9  # Standard working hours is 9
                        status = 'Present (Logs Found)'
                    else:
                        # No attendance record or logs
                        # Check for day off: Friday (4) and Saturday (5) only
                        if current_date.weekday() in [4, 5]:  # Friday/Saturday only
                            status = 'Day Off'
                        elif leave_request:
                            # FIX: Show leave type instead of Absent for approved leave days
                            if leave_request.leave_type:
                                status = leave_request.leave_type.name
                            else:
                                status = 'Annual Leave'  # Default if no type specified
                        else:
                            # Sunday (6) and Monday-Thursday (0-3) are working days and can be absent
                            # Only Friday (4) and Saturday (5) are day off
                            if current_date.weekday() in [0, 1, 2, 3, 6]:  # Monday-Thursday and Sunday
                                status = 'Absent'
                            else:
                                status = 'Day Off'  # Friday/Saturday fallback
                        hours_worked = 0.0
                        extra_time = 0.0

                # Prepare permission request info for this date
                permission_info = None
                if permission_request:
                    permission_info = {
                        'id': permission_request.id,
                        'start_time': permission_request.start_time.strftime('%H:%M:%S'),
                        'end_time': permission_request.end_time.strftime('%H:%M:%S'),
                        'duration_hours': round((permission_request.end_time - permission_request.start_time).total_seconds() / 3600, 2),
                        'reason': permission_request.reason,
                        'status': permission_request.status
                    }

                # Prepare data for this date
                # Format worked hours to HH:MM
                total_seconds = int(hours_worked * 3600)
                h, remainder = divmod(total_seconds, 3600)
                m, _ = divmod(remainder, 60)
                formatted_hours_worked = f"{h:02}:{m:02}" if (h > 0 or m > 0) else "-"

                # Format extra_time as hours and minutes
                extra_time_formatted = format_hours_minutes(extra_time)
                
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day_of_week': current_date.strftime('%A'),
                    'status': status,
                    'check_in': check_in.strftime('%H:%M:%S') if check_in else None,
                    'check_out': check_out.strftime('%H:%M:%S') if check_out else None,
                    'hours_worked': formatted_hours_worked,  # Use formatted hours
                    'raw_hours_worked': round(hours_worked, 1), # Keep raw for calculations if needed
                    'extra_time': extra_time,  # Keep raw decimal for calculations
                    'extra_time_formatted': extra_time_formatted,  # Formatted as "Xh Ym"
                    'logs': daily_logs,  # Send logs with correct structure (has 'time' and 'scan_type')
                    'all_logs': daily_logs,  # Keep for backward compatibility
                    'logs_count': len(daily_logs),
                    'permission_request': permission_info
                })
                logging.debug(f"  Appended daily data for {current_date}. Status: {status}, Hours: {hours_worked}")

                current_date += timedelta(days=1)
            except Exception as e:
                logging.error(f"Error processing date {current_date} for user {user.id}: {e}", exc_info=True)
                # Optionally append a placeholder with an error status for this day
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day_of_week': current_date.strftime('%A'),
                    'status': 'Error',
                    'check_in': None,
                    'check_out': None,
                    'hours_worked': '-',
                    'raw_hours_worked': 0.0,
                    'extra_time': 0.0,
                    'extra_time_formatted': '0h 0m',
                    'logs': [],  # Ensure logs field exists
                    'all_logs': [],  # Keep for backward compatibility
                    'logs_count': 0,
                    'permission_request': None,
                    'error_message': str(e)
                })
                current_date += timedelta(days=1) # Ensure loop progresses even on error

        return jsonify({'success': True, 'daily_data': daily_data,
                        'user_name': user.get_full_name(),
                        'total_days': len(daily_data),
                        'summary_metrics': {
                            'total_days': user_report.summary_metrics.total_days,
                            'total_working_days': user_report.summary_metrics.total_working_days,
                            'present_days': user_report.summary_metrics.present_days,
                            'absent_days': user_report.summary_metrics.absent_days,
                            'annual_leave_days': user_report.summary_metrics.annual_leave_days,
                            'paid_leave_days': user_report.summary_metrics.paid_leave_days,
                            'permission_hours': user_report.summary_metrics.permission_hours,
                            'day_off_days': user_report.summary_metrics.day_off_days,
                            'incomplete_days': user_report.summary_metrics.incomplete_days,
                            'extra_time_hours': user_report.summary_metrics.extra_time_hours
                        }
                    })

    except Exception as e:
        logging.error(f"Unhandled error in get_employee_logs for user {user_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal Server Error', 'message': str(e)}), 500
    
    except Exception as e:
        logging.error(f"Critical error in get_employee_logs for user {user_id}: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch employee logs'}), 500

@final_report_bp.route('/detailed-attendance-report/export')
@login_required
@role_required(['admin', 'director', 'support', 'product_owner', 'manager'])
def export_detailed_attendance_report():
    """Export detailed attendance report to Excel"""
    
    # Get the same parameters as the detailed attendance report
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    
    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Start date and end date are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    # Get users for export based on role
    if current_user.role == 'manager':
        # Managers can export their employees AND themselves
        from helpers import get_employees_for_manager
        team_members = get_employees_for_manager(current_user.id)
        # Include manager themselves
        users = [current_user] + list(team_members)
        # Filter to active users only
        users = [u for u in users if u.status == 'active' and 
                 not u.first_name.startswith('User') and 
                 not u.first_name.startswith('NN-') and
                 u.first_name != '' and u.last_name != '']
    else:
        # Admin, Director, Support, Product Owner can export all users
        users = User.query.filter(
            User.status == 'active',
            ~User.first_name.like('User%'),  # Exclude generic test users
            ~User.first_name.like('NN-%'),   # Exclude numbered test users
            User.first_name != '',           # Exclude empty names
            User.last_name != ''             # Exclude users without last names
        ).all()
    
    # Filter users if specific users are selected
    if user_ids:
        # For managers, ensure they can only export their own employees and themselves
        if current_user.role == 'manager':
            manager_employee_ids = [u.id for u in users]
            # Always allow manager to export themselves
            if current_user.id not in manager_employee_ids:
                manager_employee_ids.append(current_user.id)
            user_ids = [uid for uid in user_ids if uid in manager_employee_ids]
        users = [user for user in users if user.id in user_ids]
    
    # Sort users by fingerprint number: Not Assigned (None/empty) first, then numeric low to high
    def get_fingerprint_sort_key(user):
        fp = user.fingerprint_number
        if not fp or fp.strip() == '':
            return (0, 0)  # Not Assigned comes first
        try:
            fp_int = int(fp.strip())
            return (1, fp_int)  # Numeric sorting
        except (ValueError, TypeError):
            return (2, 0)  # Invalid/non-numeric comes last
    
    users = sorted(users, key=get_fingerprint_sort_key)
    
    # Generate report data using the unified calculation logic
    all_user_reports = []
    
    for user in users:
        # Use the unified calculation function to ensure exact same logic as Final Report web view
        user_report = calculate_unified_report_data(user, start_date, end_date)
        all_user_reports.append(user_report)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Name", "Fingerprint Number", "Department", "Total Days", 
        "Total Working Days", "Day Off", "Present Days", "Absent Days", 
        "Annual Leave", "Unpaid Leave", "Paid Leave", "Permission Hours", 
        "Incomplete Days", "Extra Time (hours)", "Attendance %"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Write data
    for row, user_report in enumerate(all_user_reports, 2):
        user = user_report.user
        metrics = user_report.summary_metrics
        
        ws.cell(row=row, column=1, value=user.get_full_name()).border = border
        ws.cell(row=row, column=2, value=user.fingerprint_number or 'N/A').border = border
        ws.cell(row=row, column=3, value=user.department.department_name if user.department else 'No Department').border = border
        ws.cell(row=row, column=4, value=metrics.total_days).border = border
        ws.cell(row=row, column=5, value=metrics.total_working_days).border = border
        ws.cell(row=row, column=6, value=metrics.day_off_days).border = border
        ws.cell(row=row, column=7, value=metrics.present_days).border = border
        ws.cell(row=row, column=8, value=metrics.absent_days).border = border
        ws.cell(row=row, column=9, value=metrics.annual_leave_days).border = border
        ws.cell(row=row, column=10, value=metrics.unpaid_leave_days).border = border
        ws.cell(row=row, column=11, value=metrics.paid_leave_days).border = border
        ws.cell(row=row, column=12, value=metrics.permission_hours).border = border
        ws.cell(row=row, column=13, value=metrics.incomplete_days).border = border
        # Format extra time as hours and minutes for Excel
        extra_time_formatted = format_hours_minutes(metrics.extra_time_hours)
        ws.cell(row=row, column=14, value=extra_time_formatted).border = border
        ws.cell(row=row, column=15, value=f"{metrics.attendance_percentage}%").border = border
    
    # Add summary row
    if all_user_reports:
        summary_row = len(all_user_reports) + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value="").border = border
        ws.cell(row=summary_row, column=3, value="").border = border
        ws.cell(row=summary_row, column=4, value=sum(r.summary_metrics.total_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=5, value=sum(r.summary_metrics.total_working_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=6, value=sum(r.summary_metrics.day_off_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=7, value=sum(r.summary_metrics.present_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=8, value=sum(r.summary_metrics.absent_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=9, value=sum(r.summary_metrics.annual_leave_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=10, value=sum(r.summary_metrics.unpaid_leave_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=11, value=sum(r.summary_metrics.paid_leave_days for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=12, value=sum(r.summary_metrics.permission_hours for r in all_user_reports)).border = border
        ws.cell(row=summary_row, column=13, value=sum(r.summary_metrics.incomplete_days for r in all_user_reports)).border = border
        # Format total extra time as hours and minutes
        total_extra_time = sum(r.summary_metrics.extra_time_hours for r in all_user_reports)
        total_extra_time_formatted = format_hours_minutes(total_extra_time)
        ws.cell(row=summary_row, column=14, value=total_extra_time_formatted).border = border
        ws.cell(row=summary_row, column=15, value="").border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"Detailed_Attendance_Report_{start_date_str}_to_{end_date_str}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

