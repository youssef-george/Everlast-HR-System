from flask import Blueprint, render_template, jsonify, request, redirect, url_for, make_response, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from models import LeaveRequest, PermissionRequest, User, Department, AttendanceLog
from helpers import role_required
from helpers.report_calculations import calculate_multiple_users_report_data
from sqlalchemy import func, or_, and_
import io
import os

calendar_bp = Blueprint('calendar', __name__, url_prefix='/calendar')

@calendar_bp.route('/')
@login_required
def index():
    from models import User
    import threading
    import logging
    from connection_manager import is_sync_running
    
    # Auto-sync data from device when page loads (non-blocking) - only if no sync is running
    if not is_sync_running():
        try:
            from routes.attendance import sync_attendance_task
            def sync_task():
                try:
                    sync_attendance_task(full_sync=True)
                except Exception as e:
                    logging.error(f'Error auto-syncing data on calendar main page load: {str(e)}')
            
            sync_thread = threading.Thread(target=sync_task, daemon=True)
            sync_thread.start()
        except Exception as e:
            logging.error(f'Error starting sync thread for calendar main page: {str(e)}')
    else:
        logging.info('Skipping sync on calendar page load - another sync is already running')
    
    # Build user list for filter based on role
    if current_user.role == 'manager':
        from helpers import get_employees_for_manager
        employees = get_employees_for_manager(current_user.id)
        all_users = [emp for emp in employees if emp.status == 'active'] + [current_user]
    elif current_user.role in ['admin', 'product_owner', 'director']:
        all_users = User.query.filter_by(status='active').order_by(User.first_name).all()
    else:
        # Employees do not need the filter; pass only self for completeness
        all_users = [current_user] if current_user.status == 'active' else []
    
    """Show the calendar page"""
    return render_template('calendar/index.html', title='Calendar', all_users=all_users)

@calendar_bp.route('/events')
@login_required
def events():
    """API endpoint to get calendar events based on user role"""
    try:
        import logging
        from flask import g
        
        logging.info(f"Calendar events request started for user {current_user.id} with role {current_user.role}")
        
        # Simple test to see if basic functionality works
        if request.args.get('test') == 'true':
            return jsonify({'test': 'success', 'user_id': current_user.id, 'role': current_user.role})
        
        # Simple caching to prevent repeated processing
        cache_key = f"calendar_events_{current_user.id}_{request.args.get('user_id', 'all')}"
        if hasattr(g, cache_key):
            return jsonify(g.cache_key)
            
        filter_user_id = request.args.get('user_id', type=int)
        user_role = current_user.role
        logging.info(f"Calendar events request - User: {current_user.id}, Role: {user_role}, Filter User ID: {filter_user_id}")
        # Extended date range to show all past days and future data
        start_date = datetime.today().date() - timedelta(days=365*5)   # Show from 5 years ago
        end_date = datetime.today().date() + timedelta(days=365*2)    # Show 2 years ahead
        
        events = []
        
        if user_role == 'employee':
            # Employees see only their own leaves and permissions
            user_id = current_user.id
            if filter_user_id and filter_user_id != user_id:
                # Employees cannot view other users' data
                logging.debug("Employee trying to view other user's data. Returning empty.")
                return jsonify([])
        elif filter_user_id:
            # If a specific user is selected, override role-based filtering for managers/admins
            user_id = filter_user_id
            # Check if the current user is authorized to view this specific user's data
            if user_role == 'manager':
                from helpers import get_employees_for_manager
                managed_employees = get_employees_for_manager(current_user.id)
                if user_id not in [emp.id for emp in managed_employees] and user_id != current_user.id:
                    logging.debug("Manager trying to view unauthorized user's data. Returning empty.")
                    return jsonify([])
            elif user_role in ['admin', 'product_owner', 'director']:
                # Admins/Product Owners/Directors can view any user's data, but we still set user_id for specific filtering
                pass
        else:
            # No specific user filter - show data based on role
            if user_role == 'manager':
                from helpers import get_employees_for_manager
                managed_employees = get_employees_for_manager(current_user.id)
                user_id = None  # Will be handled in the query
            elif user_role in ['admin', 'product_owner', 'director']:
                user_id = None  # Will show all users
            else:
                user_id = current_user.id
            
        # Get users based on role and filter
        try:
            if user_id:
                # Specific user selected - only show that user's events
                users = User.query.filter_by(id=user_id, status='active').all()
                logging.info(f"Filtering to specific user: {user_id}")
            elif user_role == 'manager':
                # Manager sees their department employees
                from helpers import get_employees_for_manager
                employees = get_employees_for_manager(current_user.id)
                users = [emp for emp in employees if emp.status == 'active'] + [current_user]
                logging.info(f"Manager view - showing {len(users)} users")
            elif user_role in ['admin', 'product_owner', 'director']:
                # Admin/Product Owner/Director sees all active users
                users = User.query.filter_by(status='active').all()
                logging.info(f"Admin/Product Owner/Director view - showing {len(users)} users")
            else:
                # Employee sees only themselves
                users = [current_user] if current_user.status == 'active' else []
                logging.info(f"Employee view - showing {len(users)} users")
        except Exception as e:
            logging.error(f"Error getting users: {str(e)}")
            raise
        
        logging.info(f"Processing {len(users)} users for calendar events")
        
        # Get leaves and permissions for the date range
        try:
            if user_id:
                leaves = LeaveRequest.query.filter(
                    LeaveRequest.user_id == user_id,
                    LeaveRequest.start_date <= end_date,
                    LeaveRequest.end_date >= start_date
                ).all()
                permissions = PermissionRequest.query.filter(
                    PermissionRequest.user_id == user_id,
                    func.date(PermissionRequest.start_time) >= start_date,
                    func.date(PermissionRequest.start_time) <= end_date
                ).all()
            else:
                # Get leaves and permissions for all users
                user_ids = [user.id for user in users]
                leaves = LeaveRequest.query.filter(
                    LeaveRequest.user_id.in_(user_ids),
                    LeaveRequest.start_date <= end_date,
                    LeaveRequest.end_date >= start_date
                ).all()
                permissions = PermissionRequest.query.filter(
                    PermissionRequest.user_id.in_(user_ids),
                    func.date(PermissionRequest.start_time) >= start_date,
                    func.date(PermissionRequest.start_time) <= end_date
                ).all()
            
            logging.info(f"Found {len(leaves)} leave requests and {len(permissions)} permission requests")
        except Exception as e:
            logging.error(f"Error getting leaves and permissions: {str(e)}")
            raise
        
        # Add leave events
        for leave in leaves:
            color = '#17a74a' if leave.status == 'approved' else '#dc3545' if leave.status == 'rejected' else '#ffc107'
            events.append({
                'id': f"leave_{leave.id}",
                'title': f"{leave.user.get_full_name()} - Leave",
                'start': leave.start_date.isoformat(),
                'end': (leave.end_date + timedelta(days=1)).isoformat(),
                'color': color,
                'url': f"/leave/view/{leave.id}",
                'status': leave.status,
                'type': 'leave'
            })
        
        # Add permission events
        for permission in permissions:
            color = '#17a74a' if permission.status == 'approved' else '#dc3545' if permission.status == 'rejected' else '#ffc107'
            events.append({
                'id': f"permission_{permission.id}",
                'title': f"{permission.user.get_full_name()} - Permission",
                'start': permission.start_time.isoformat(),
                'end': permission.end_time.isoformat(),
                'color': color,
                'url': f"/permission/view/{permission.id}",
                'status': permission.status,
                'type': 'permission'
            })
        
        # Add paid holiday events
        try:
            from models import PaidHoliday
            paid_holidays = PaidHoliday.query.filter(
                PaidHoliday.start_date >= start_date,
                PaidHoliday.start_date <= end_date
            ).all()
            
            for holiday in paid_holidays:
                if holiday.holiday_type == 'day':
                    # Single day holiday
                    events.append({
                        'id': f"holiday_{holiday.id}",
                        'title': f"🏖️ {holiday.description}",
                        'start': holiday.start_date.isoformat(),
                        'end': (holiday.start_date + timedelta(days=1)).isoformat(),
                        'color': '#6f42c1',  # Purple color for holidays
                        'url': f"/dashboard/paid-holidays",
                        'status': 'holiday',
                        'type': 'holiday',
                        'allDay': True
                    })
                else:
                    # Range holiday
                    events.append({
                        'id': f"holiday_{holiday.id}",
                        'title': f"🏖️ {holiday.description}",
                        'start': holiday.start_date.isoformat(),
                        'end': (holiday.end_date + timedelta(days=1)).isoformat(),
                        'color': '#6f42c1',  # Purple color for holidays
                        'url': f"/dashboard/paid-holidays",
                        'status': 'holiday',
                        'type': 'holiday',
                        'allDay': True
                    })
        except Exception as e:
            logging.error(f"Error getting paid holidays: {str(e)}")
        
        # Process attendance only for past dates and today (not future dates)
        try:
            today = datetime.today().date()
            current_date = start_date
            while current_date <= end_date:
                # Skip future dates for attendance processing
                if current_date > today:
                    current_date += timedelta(days=1)
                    continue
                    
                # Get attendance logs for this date
                start_datetime = datetime.combine(current_date, datetime.min.time())
                end_datetime = datetime.combine(current_date, datetime.max.time())
                
                # Optimize query - only get logs for users we care about
                user_ids_list = [user.id for user in users]
                today_logs = AttendanceLog.query.filter(
                    AttendanceLog.timestamp.between(start_datetime, end_datetime),
                    AttendanceLog.user_id.in_(user_ids_list)
                ).order_by(AttendanceLog.timestamp.desc()).all()
            
                # Process logs using same logic as Daily Attendance
                processed_logs = {}
                for log in today_logs:
                    if log.user_id not in processed_logs:
                        processed_logs[log.user_id] = {
                            'user': log.user,
                            'check_in': None,
                            'check_out': None,
                            'duration': None,
                            'status': 'absent',
                            'all_logs': []  # Store all logs for this user
                        }
                    
                    # Add log to all_logs
                    processed_logs[log.user_id]['all_logs'].append(log)
                    
                    # Determine check-in/check-out using same logic as Daily Attendance
                    log_type = 'check-in' if log.timestamp.hour < 12 else 'check-out'
                    if log_type == 'check-in' and (not processed_logs[log.user_id]['check_in'] or 
                                                  log.timestamp < processed_logs[log.user_id]['check_in'].timestamp):
                        processed_logs[log.user_id]['check_in'] = log
                        processed_logs[log.user_id]['status'] = 'in_office'
                    elif log_type == 'check-out' and (not processed_logs[log.user_id]['check_out'] or 
                                                     log.timestamp > processed_logs[log.user_id]['check_out'].timestamp):
                        processed_logs[log.user_id]['check_out'] = log
                    
                    # Calculate duration and update status if both exist
                    if processed_logs[log.user_id]['check_in'] and processed_logs[log.user_id]['check_out']:
                        duration = processed_logs[log.user_id]['check_out'].timestamp - processed_logs[log.user_id]['check_in'].timestamp
                        processed_logs[log.user_id]['duration'] = duration
                        processed_logs[log.user_id]['status'] = 'present'
                
                # Sort all_logs by timestamp for each user
                for user_id_key, data in processed_logs.items():
                    data['all_logs'].sort(key=lambda x: x.timestamp)
                
                # Only show users who have actual attendance records, leave, or permission for this date
                for user in users:
                    is_weekend = current_date.weekday() in [4, 5]  # Friday/Saturday
                    
                    # Check if before joining date
                    if user.joining_date and current_date < user.joining_date:
                        continue  # Don't show users before their joining date
                    
                    # Don't show attendance data for future dates
                    if current_date > today:
                        continue  # Skip future dates
                    
                    # Check for leave
                    leave_request = LeaveRequest.query.filter(
                        LeaveRequest.user_id == user.id,
                        LeaveRequest.status == 'approved',
                        LeaveRequest.start_date <= current_date,
                        LeaveRequest.end_date >= current_date
                    ).first()
                    
                    # Check for permission
                    permission_request = PermissionRequest.query.filter(
                        PermissionRequest.user_id == user.id,
                        func.date(PermissionRequest.start_time) == current_date,
                        PermissionRequest.status == 'approved'
                    ).first()
                    
                    # Only show user if they have attendance, leave, or permission records
                    should_show_user = False
                    status = ''
                    color = '#6c757d'
                    
                    if user.id in processed_logs:
                        # User has attendance logs
                        user_data = processed_logs[user.id]
                        should_show_user = True
                        
                        if leave_request:
                            # Get specific leave type name
                            leave_type_name = leave_request.leave_type.name if leave_request.leave_type else 'Leave Request'
                            status = leave_type_name
                            color = '#ffc107'  # Yellow for leave
                        elif permission_request:
                            status = 'Permission'
                            color = '#17a2b8'  # Blue for permission
                        elif is_weekend and user_data['status'] == 'present':
                            status = 'Day Off / Present'
                            color = '#28a745'  # Green for present on weekend
                        elif user_data['status'] == 'present':
                            status = 'Present'
                            color = '#28a745'  # Green for present
                        elif user_data['status'] == 'in_office':
                            status = 'In Office'
                            color = '#6f42c1'  # Purple for in office
                        else:
                            status = 'Present'  # Default to present if has logs
                            color = '#28a745'
                            
                    elif leave_request:
                        # User has leave but no attendance
                        should_show_user = True
                        # Get specific leave type name
                        leave_type_name = leave_request.leave_type.name if leave_request.leave_type else 'Leave Request'
                        status = leave_type_name
                        color = '#ffc107'  # Yellow for leave
                        
                    elif permission_request:
                        # User has permission but no attendance
                        should_show_user = True
                        status = 'Permission'
                        color = '#17a2b8'  # Blue for permission
                        
                    elif not is_weekend:
                        # User is absent on a weekday (only show absent users on weekdays)
                        should_show_user = True
                        status = 'Absent'
                        color = '#dc3545'  # Red for absent
                    
                    # Only add event if user should be shown
                    if should_show_user:
                        events.append({
                            'id': f"attendance_{user.id}_{current_date}",
                            'title': f"{user.get_full_name()} - {status}",
                            'start': current_date.isoformat(),
                            'end': current_date.isoformat(),
                            'color': color,
                            'url': f"/attendance/",
                            'status': status,
                            'type': 'attendance'
                        })
                
                current_date += timedelta(days=1)
        except Exception as e:
            logging.error(f"Error processing attendance logs: {str(e)}")
            raise
        
        # Store in cache for this request
        setattr(g, cache_key, events)
        
        logging.info(f"Returning {len(events)} events for calendar")
        if filter_user_id:
            logging.info(f"Events for user {filter_user_id}: {[event.get('title', 'No title') for event in events[:5]]}")
        
        return jsonify(events)
    
    except Exception as e:
        import traceback
        logging.error(f"Error in events(): {str(e)}")
        logging.error(f"Traceback: {traceback.format_exc()}")
        
        # Return empty events array instead of error to prevent calendar from breaking
        return jsonify([])


@calendar_bp.route('/attendance-report')
@login_required
def attendance_report():
    """Calendar Attendance Report - uses same logic as Daily Attendance page"""
    # Redirect employees to their personal attendance page
    if current_user.role == 'employee':
        return redirect(url_for('attendance.my_attendance'))
    
    # Clean up orphaned paid holiday records before processing
    from routes.attendance import cleanup_orphaned_paid_holiday_records
    cleanup_orphaned_paid_holiday_records()
    
    from models import DailyAttendance, LeaveRequest, PermissionRequest, AttendanceLog, PaidHoliday
    from datetime import date
    import threading
    import logging
    from connection_manager import is_sync_running
    
    # Auto-sync disabled - manual sync only
    # if not is_sync_running():
    #     try:
    #         from routes.attendance import sync_attendance_task
    #         def sync_task():
    #             try:
    #                 sync_attendance_task(full_sync=True)
    #             except Exception as e:
    #                 logging.error(f'Error auto-syncing data on calendar attendance-report page load: {str(e)}')
    #         
    #         sync_thread = threading.Thread(target=sync_task, daemon=True)
    #         sync_thread.start()
    #     except Exception as e:
    #         logging.error(f'Error starting sync thread for calendar attendance-report: {str(e)}')
    # else:
    #     logging.info('Skipping sync on calendar attendance-report page load - another sync is already running')
    
    # Get date range from query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    
    # Get users for dropdown based on role (always needed for the dropdown)
    if current_user.role == 'manager':
        # Manager sees their department employees
        from helpers import get_employees_for_manager
        employees = get_employees_for_manager(current_user.id)
        users = [emp for emp in employees if emp.status == 'active'] + [current_user]
    elif current_user.role in ['admin', 'product_owner', 'director']:
        # Admin/Product Owner/Director sees all active users
        users = User.query.filter_by(status='active').all()
    else:
        # Employee sees only themselves
        users = [current_user] if current_user.status == 'active' else []
    
    # Don't show any data by default - only show when dates are explicitly provided
    if not start_date_str or not end_date_str:
        # Return empty report if no dates provided
        return render_template('calendar/attendance_report.html', 
                             user_reports=[], 
                             start_date=None, 
                             end_date=None,
                             total_users=0,
                             users=users)
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Limit date range to prevent extremely long processing times
    days_diff = (end_date - start_date).days + 1
    if days_diff > 365:  # More than 1 year
        return render_template('calendar/attendance_report.html', 
                             user_reports=[], 
                             start_date=start_date, 
                             end_date=end_date,
                             total_users=0,
                             users=users,
                             error_message="Date range cannot exceed 365 days. Please select a smaller range.")
    
    # Get sorting preference
    newest_first = request.args.get('newest_first', '').strip().lower() == 'on'
    
    # Filter users if specific users are selected
    if user_ids:
        # Specific users selected - filter to only those users
        users = [user for user in users if user.id in user_ids]
    
    # Optimize: Fetch all attendance logs for all users in the date range at once
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    user_ids = [user.id for user in users]
    logging.info(f"Fetching attendance logs for {len(users)} users from {start_date} to {end_date}")
    
    all_attendance_logs = AttendanceLog.query.filter(
        AttendanceLog.user_id.in_(user_ids),
        AttendanceLog.timestamp.between(start_datetime, end_datetime)
    ).order_by(AttendanceLog.user_id, AttendanceLog.timestamp).all()
    
    logging.info(f"Retrieved {len(all_attendance_logs)} attendance logs")
    
    # Group logs by user and date for efficient processing
    logs_by_user_date = {}
    for log in all_attendance_logs:
        log_date = log.timestamp.date()
        if log.user_id not in logs_by_user_date:
            logs_by_user_date[log.user_id] = {}
        if log_date not in logs_by_user_date[log.user_id]:
            logs_by_user_date[log.user_id][log_date] = []
        logs_by_user_date[log.user_id][log_date].append(log)
    
    # Generate report data for each user using shared calculation logic
    all_user_reports = calculate_multiple_users_report_data(users, start_date, end_date)
    
    # Convert to the format expected by the template
    formatted_user_reports = []
    for user_report in all_user_reports:
        # Create a compatible structure for the template
        formatted_report = {
            'user': user_report.user,
            'summary_metrics': user_report.summary_metrics,
            'report_data': user_report.report_data
        }
        formatted_user_reports.append(formatted_report)
    
    return render_template('calendar/attendance_report.html',
                         users=users, 
                         start_date=start_date, 
                         end_date=end_date, 
                         all_user_reports=formatted_user_reports,
                         newest_first=newest_first)

@calendar_bp.route('/summary')
@login_required
def summary():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_id = request.args.get('user_id')

    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Start date and end date are required'}), 400

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # If a specific user is selected, use that user_id, otherwise use current_user.id
    target_user_id = int(user_id) if user_id and user_id.isdigit() else current_user.id

    from models import DailyAttendance, LeaveRequest # Import DailyAttendance model

    present_days = 0
    absent_days = 0
    day_offs = 0
    leave_days = 0
    effective_days = 0
    extra_hours = 0.0

    # Fetch daily attendance records for the user within the date range
    attendance_records = DailyAttendance.query.filter(
        DailyAttendance.user_id == target_user_id,
        DailyAttendance.date >= start_date,
        DailyAttendance.date <= end_date
    ).all()

    # Fetch approved leave requests for the user within the date range
    approved_leaves = LeaveRequest.query.filter(
        LeaveRequest.user_id == target_user_id,
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= end_date,
        LeaveRequest.end_date >= start_date
    ).all()

    # Calculate metrics
    for record in attendance_records:
        if record.status == 'present':
            present_days += 1
            effective_days += 1
            if record.total_hours and record.total_hours > 8:
                extra_hours += record.total_hours - 8
        elif record.status == 'absent':
            absent_days += 1
        elif record.status == 'day_off':
            day_offs += 1

    # Count leave days
    for leave in approved_leaves:
        # Calculate overlap with the selected date range
        overlap_start = max(leave.start_date, start_date)
        overlap_end = min(leave.end_date, end_date)
        if overlap_start <= overlap_end:
            leave_days += (overlap_end - overlap_start).days + 1

    return jsonify({
        'present_days': present_days,
        'absent_days': absent_days,
        'day_offs': day_offs,
        'leave_days': leave_days,
        'effective_days': effective_days,
        'extra_hours': round(extra_hours, 2)
    })

@calendar_bp.route('/export-attendance-report')
@login_required
@role_required(['admin', 'product_owner'])
def export_attendance_report():
    """Export attendance report to Excel or PDF"""
    
    # Get the same parameters as the attendance report
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    user_ids = request.args.getlist('user_ids', type=int)
    export_format = request.args.get('format', 'excel').lower()  # excel or pdf
    
    if not start_date_str or not end_date_str:
        return jsonify({'error': 'Start date and end date are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    # Get users for export based on role
    if current_user.role == 'manager':
        from helpers import get_employees_for_manager
        employees = get_employees_for_manager(current_user.id)
        users = [emp for emp in employees if emp.status == 'active'] + [current_user]
    elif current_user.role in ['admin', 'product_owner', 'director']:
        users = User.query.filter_by(status='active').all()
    else:
        users = [current_user] if current_user.status == 'active' else []
    
    # Filter users if specific users are selected
    if user_ids:
        users = [user for user in users if user.id in user_ids]
    
    # Calculate report data using shared calculation logic
    all_user_reports = calculate_multiple_users_report_data(users, start_date, end_date)
    
    # Export based on format
    if export_format == 'pdf':
        return export_to_pdf(all_user_reports, start_date, end_date)
    else:
        return export_to_excel(all_user_reports, start_date, end_date)

def export_to_excel(all_user_reports, start_date, end_date):
    """Export attendance report to Excel format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'Excel export not available. Please install openpyxl package.'}), 500
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title = f"Attendance Report ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ['Employee', 'Present Days', 'Absent Days', 'Leave Days', 'Day Off', 'Extra Hours', 'Total Days', 'Effective Days']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    row = 4
    for user_report in all_user_reports:
        user = user_report['user']
        metrics = user_report['summary_metrics']
        
        ws.cell(row=row, column=1, value=f"{user.first_name} {user.last_name}").border = border
        ws.cell(row=row, column=2, value=metrics['present_days']).border = border
        ws.cell(row=row, column=3, value=metrics['absent_days']).border = border
        ws.cell(row=row, column=4, value=metrics['annual_leave_days'] + metrics['paid_leave_days']).border = border
        ws.cell(row=row, column=5, value=metrics['day_off_days']).border = border
        ws.cell(row=row, column=6, value=f"{metrics['extra_time_hours']:.2f}").border = border
        ws.cell(row=row, column=7, value=metrics['total_days']).border = border
        ws.cell(row=row, column=8, value=metrics['total_working_days']).border = border
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename
    filename = f"Attendance_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

def export_to_pdf(all_user_reports, start_date, end_date):
    """Export attendance report to PDF format"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import inch
    except ImportError:
        return jsonify({'error': 'PDF export not available. Please install reportlab package.'}), 500
    
    output = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Add title
    title = f"Attendance Report ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 12))
    
    # Create summary table
    table_data = [['Employee', 'Present', 'Absent', 'Leave', 'Day Off', 'Extra Hours', 'Total', 'Effective']]
    
    for user_report in all_user_reports:
        user = user_report['user']
        metrics = user_report['summary_metrics']
        
        table_data.append([
            f"{user.first_name} {user.last_name}",
            str(metrics['present_days']),
            str(metrics['absent_days']),
            str(metrics['annual_leave_days'] + metrics['paid_leave_days']),
            str(metrics['day_off_days']),
            f"{metrics['extra_time_hours']:.2f}",
            str(metrics['total_days']),
            str(metrics['total_working_days'])
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    # Create filename
    filename = f"Attendance_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.pdf"
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
